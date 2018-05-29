#-*- coding: UTF-8 -*-
import urllib.request
from openpyxl import Workbook
import re


def pretend():
    headers = {
        'Accept':
        'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Encoding':
        'gb2312,utf-8',
        'Accept-Language':
        'zh-CN,zh;q=0.8',
        'Connection':
        'keep-alive',
        'User-Agent':
        'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.235',
    }
    opener = urllib.request.build_opener()
    headall = []
    for key, value in headers.items():
        item = (key, value)
        headall.append(item)
    opener.addheaders = headall
    urllib.request.install_opener(opener)


def get_books(book_tag):
    url = 'https://book.douban.com/tag/' + urllib.request.quote(
        book_tag) + '?start='

    bookPat = re.compile('<li class="subject-item">(.+?)</li>', re.S)
    namePat = re.compile(
        '<a href="https://book.douban.com/subject/\d+?/" title="(.+?)"')
    ratingPat = re.compile('<span class="rating_nums">(.+?)</span>')
    peoplePat = re.compile('<span class="pl">.+?\((.+?)人', re.S)
    infoPat = re.compile(
        '<div class="pub">\n        \n  \n  (.*?)\n\n      </div>')
    book_info = []
    for i in range(100):
        try:
            u = url + str(i * 20)
            print(u)
            r = urllib.request.urlopen(u).read().decode(
                'utf-8', errors='ignore')
            book = bookPat.findall(r)
            if book == []:
                break
            for i in book:
                name = namePat.findall(i)[0]
                rating = ratingPat.findall(i)
                if rating == []:
                    rating = ''
                else:
                    rating = rating[0]
                people = peoplePat.findall(i)
                if people == []:
                    people = ''
                else:
                    people = people[0]
                info = infoPat.findall(i)
                if info == []:
                    auther = '暂无'
                    publishing = ''
                    time = ''
                    money = ''

                else:
                    try:
                        info_list = info[0].split('/')
                        auther = '/'.join(info_list[:-3])
                        publishing = info_list[-3]
                        time = info_list[-2]
                        money = info_list[-1]
                    except:
                        auther = info[0]
                        publishing = ''
                        time = ''
                        money = ''

                book_info.append([name, rating, people, auther, publishing, time, money])
        except:
            pass
    return book_info


def get_all_books(book_tag_list):
    all_book_list = []
    for book_tag in book_tag_list:
        all_book_list.append(get_books(book_tag))
    return all_book_list


def get_book_list_excel(book_tag_list, all_book_list):
    wb = Workbook()
    ws = []
    for i in range(len(book_tag_list)):
        ws.append(wb.create_sheet(book_tag_list[i]))
    for i in range(len(book_tag_list)):
        ws[i].append(['书名', '评分', '评价人数', '作者/译者','出版社','出版时间','价钱'])
        for book in all_book_list[i]:
            ws[i].append([book[0], book[1], book[2], book[3],book[4], book[5], book[6]])

    mark = '-'
    file_name = mark.join(book_tag_list) + '.xlsx'
    wb.save(file_name)


def main():
    # book_tag_list = ['小说','随笔','名著','童话']
    # book_tag_list = ['漫画','推理','科幻','奇幻','武侠']
    # book_tag_list = ['经济学','管理','商业','金融','理财']
    book_tag_list = ['科普', '互联网', '神经网络', '算法']
    
    pretend()
    all_book_list = get_all_books(book_tag_list)
    get_book_list_excel(book_tag_list, all_book_list)

    print("Done!")


if __name__ == '__main__':
    main()