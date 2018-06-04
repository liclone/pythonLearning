import pymysql
import datetime
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook


def get_rank(items):
	rank = []
	conn = pymysql.connect(
		host='127.0.0.1',
    	user='root',
    	passwd='666666',
    	db='bilibiliVideoInfo',
    	charset='utf8')
	cursor = conn.cursor()
	dt = datetime.datetime.now()
	table_name = dt.strftime('%Y_%m_%d')
	for item in items:
		cursor.execute('select * from {table_name} order by {item} desc limit 0,20'.format(table_name=table_name, item=item))
		rank.append(cursor.fetchall())

	return rank

def get_excel(items, rank):
	url_videoName = 'https://www.bilibili.com/video/av{aid}'
	headers={
	'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
	'Accept-Encoding': 'gb2312,utf-8',
	'Accept-Language': 'zh-CN,zh;q=0.8',
	'Connection': 'keep-alive',
	'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/43.0.235',
	}

	wb = Workbook()
	ws = []
	
	for item in items:
		ws.append(wb.create_sheet(item))
	for i in range(len(items)):
		ws[i].append(['No.','aid','播放','弹幕','收藏','分享数','硬币','回复','视频名'])
		j = 0
		for r in rank[i]:
			j += 1
			name = BeautifulSoup(requests.get(url_videoName.format(aid=r[0]),headers = headers).text,'html.parser').head.title.text.strip('_哔哩哔哩 (゜-゜)つロ 干杯~-bilibili ')
			ws[i].append([j,r[0],r[1],r[2],r[3],r[4],r[5],r[6],name])

	dt = datetime.datetime.now()
	wb.save(dt.strftime('%Y_%m_%d') + '.xlsx')

	print("Done!")



def main():
	items = ['view','danmaku','favorite','share','coin','reply']
	rank = get_rank(items)
	get_excel(items,rank)


if __name__ == "__main__":
	main()
